#!/usr/bin/env python3
"""Export data/ingredients.xlsx to JSON files Godot loads at runtime."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "ingredients.xlsx"
INGREDIENTS_JSON = ROOT / "data" / "ingredients.json"
STARTER_BAG_JSON = ROOT / "data" / "starter_bag.json"
AURAS_JSON = ROOT / "data" / "auras.json"
TRINKETS_JSON = ROOT / "data" / "trinkets.json"

INGREDIENT_HEADERS = [
    "id",
    "art",
    "display_name",
    "description",
    "point_value",
    "explosive_value",
    "shop_cost",
    "rarity",
    "shop_available",
]

STARTER_HEADERS = ["ingredient_id", "count"]

AURA_HEADERS = [
    "id",
    "display_name",
    "description",
    "pool",
    "pool_unlock_level",
    "explosion_limit_modifier",
    "score_multiplier_percent",
    "gold_multiplier_percent",
]

TRINKET_HEADERS = [
    "id",
    "display_name",
    "description",
    "reward_offerable",
]

DEFAULT_INGREDIENTS = [
    ["boom_berry_1", "boom_berry", "Small Boom Berry", "A tiny berry that pops.", 1, 1, 4, "common", False],
    ["boom_berry_2", "boom_berry", "Boom Berry", "A ripe berry with a bigger blast.", 2, 2, 6, "common", False],
    ["boom_berry_3", "boom_berry", "Large Boom Berry", "An enormous berry. Handle with care.", 3, 3, 8, "uncommon", False],
]

DEFAULT_STARTER_BAG = [
    ["boom_berry_1", 4],
    ["boom_berry_2", 2],
]

DEFAULT_TRINKETS = [
    [
        "pumpkin_trinket",
        "Pumpkin Necklace",
        "Pumpkins gain +1 score for each pumpkin played before it in a row (maximum of +3)",
        True,
    ],
    [
        "red_mushroom_trinket",
        "Red Mushroom Trinket",
        "Red mushrooms can score a maximum of 6 instead of 4",
        True,
    ],
    [
        "rat_trinket",
        "Rat Trinket",
        "Rats can score a maximum of 6 instead of 4",
        True,
    ],
]

DEFAULT_AURAS = [
    [
        "gentle_brew",
        "Gentle Brew",
        "Explosion limit +1 this level.",
        "normal",
        1,
        1,
        100,
        100,
    ],
    [
        "gold_rush",
        "Gold Rush",
        "Gold earned +25% this level.",
        "normal",
        1,
        0,
        100,
        125,
    ],
    [
        "shaky_hands",
        "Shaky Hands",
        "Explosion limit -1 this level.",
        "normal",
        1,
        -1,
        100,
        100,
    ],
    [
        "overpressure",
        "Overpressure",
        "Boss: explosion limit -2.",
        "boss",
        1,
        -2,
        100,
        100,
    ],
    [
        "stingy_market",
        "Stingy Market",
        "Boss: gold earned -25%.",
        "boss",
        1,
        0,
        100,
        75,
    ],
]

INPUT_FONT = Font(name="Arial", color="0000FF")
HEADER_FONT = Font(name="Arial", bold=True)
NOTE_FONT = Font(name="Arial", italic=True, color="444444")
HEADER_FILL = PatternFill("solid", fgColor="E8E8E8")
FLAVOR_FONT = Font(name="Arial", italic=True, color="0000FF")


def _header_map(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for idx, cell in enumerate(sheet[1], start=1):
        if cell.value is None:
            continue
        headers[str(cell.value).strip()] = idx
    return headers


def _description_from_cell(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        parts: list[str] = []
        for item in value:
            if isinstance(item, TextBlock):
                text = item.text
                if item.font and item.font.i:
                    parts.append(f"[i]{text}[/i]")
                else:
                    parts.append(text)
            else:
                parts.append(str(item))
        return "".join(parts).strip()
    text = str(value).strip()
    if cell.font and cell.font.i:
        return f"[i]{text}[/i]"
    return text


def _normalize_art(value, ingredient_id: str) -> str:
    if value is None or str(value).strip() == "":
        return ingredient_id
    art = str(value).strip().replace("\\", "/")
    if "/" in art:
        art = art.rsplit("/", 1)[-1]
    if art.lower().endswith(".png"):
        art = art[:-4]
    return art


def _default_shop_available(ingredient_id: str) -> bool:
    return not str(ingredient_id).startswith("boom_berry")


def _parse_bool(value, ingredient_id: str = "", default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return _default_shop_available(ingredient_id) if ingredient_id else default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def _parse_int(value, field_name: str, row_index: int, ingredient_id: str) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError(
            f"Row {row_index} ('{ingredient_id}'): '{field_name}' is required."
        )
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Row {row_index} ('{ingredient_id}'): '{field_name}' must be a number."
        ) from exc


def _parse_int_default(
    value,
    field_name: str,
    row_index: int,
    ingredient_id: str,
    default: int = 0,
) -> int:
    if value is None or str(value).strip() == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Row {row_index} ('{ingredient_id}'): '{field_name}' must be a number."
        ) from exc


def _sheet_rows_values(sheet) -> list[list]:
    rows: list[list] = []
    for row in sheet.iter_rows(values_only=True):
        if row is None:
            continue
        cells = list(row)
        if not any(cell is not None and str(cell).strip() != "" for cell in cells):
            continue
        rows.append(cells)
    return rows


def migrate_ingredients_sheet(wb) -> bool:
    if "Ingredients" not in wb.sheetnames:
        return False
    sheet = wb["Ingredients"]
    columns = _header_map(sheet)
    if (
        set(INGREDIENT_HEADERS).issubset(columns)
        and "explosive_equals_score" not in columns
        and "bonus_score_on_draw" not in columns
    ):
        return False

    rows: list[list] = []
    for row_idx in range(2, sheet.max_row + 1):
        ingredient_id = sheet.cell(row_idx, columns["id"]).value if "id" in columns else None
        if ingredient_id is None or str(ingredient_id).strip() == "":
            continue
        desc_cell = sheet.cell(row_idx, columns["description"]) if "description" in columns else None
        description = _description_from_cell(desc_cell) if desc_cell is not None else ""
        point_value = sheet.cell(row_idx, columns["point_value"]).value if "point_value" in columns else 0
        explosive_value = (
            sheet.cell(row_idx, columns["explosive_value"]).value
            if "explosive_value" in columns
            else 0
        )
        if "explosive_equals_score" in columns:
            equals_score = sheet.cell(row_idx, columns["explosive_equals_score"]).value
            if equals_score in (True, "TRUE", "true", 1, "1"):
                explosive_value = point_value
        shop_cost = sheet.cell(row_idx, columns["shop_cost"]).value if "shop_cost" in columns else None
        if shop_cost is None:
            shop_cost = max(2, int(point_value or 0) * 2)
        rarity = sheet.cell(row_idx, columns["rarity"]).value if "rarity" in columns else None
        if rarity is None:
            rarity = "common"
        display_name = (
            sheet.cell(row_idx, columns["display_name"]).value
            if "display_name" in columns
            else ingredient_id
        )
        ingredient_id = str(ingredient_id).strip()
        if "art" in columns:
            art = _normalize_art(sheet.cell(row_idx, columns["art"]).value, ingredient_id)
        else:
            art = ingredient_id
        if "shop_available" in columns:
            shop_available = _parse_bool(
                sheet.cell(row_idx, columns["shop_available"]).value,
                ingredient_id,
            )
        else:
            shop_available = _default_shop_available(ingredient_id)
        rows.append(
            [
                ingredient_id,
                art,
                str(display_name).strip() if display_name is not None else "",
                description,
                point_value,
                explosive_value,
                shop_cost,
                str(rarity).strip().lower() if rarity is not None else "common",
                shop_available,
            ]
        )

    if "Ingredients" in wb.sheetnames:
        del wb["Ingredients"]
    ingredients = wb.create_sheet("Ingredients", 1)
    ingredients.append(INGREDIENT_HEADERS)
    for row in rows:
        ingredients.append(row)
    for cell in ingredients[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in ingredients.iter_rows(min_row=2, max_col=len(INGREDIENT_HEADERS)):
        for cell in row:
            if cell.column == 4:
                continue
            cell.font = INPUT_FONT
    rarity_validation = DataValidation(
        type="list",
        formula1='"common,uncommon,rare,epic,legendary"',
        allow_blank=False,
    )
    rarity_validation.add(f"H2:H{max(ingredients.max_row, 200)}")
    ingredients.add_data_validation(rarity_validation)
    shop_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    shop_validation.add(f"I2:I{max(ingredients.max_row, 200)}")
    ingredients.add_data_validation(shop_validation)
    widths = [18, 18, 20, 44, 12, 14, 10, 12, 14]
    for idx, width in enumerate(widths, start=1):
        ingredients.column_dimensions[chr(64 + idx)].width = width
    try:
        wb.save(XLSX_PATH)
        print(f"Migrated spreadsheet columns in {XLSX_PATH}")
    except OSError as exc:
        print(
            f"Could not save migrated Ingredients sheet ({exc}). "
            "Close ingredients.xlsx in Excel and export again to update the file."
        )
    return True


def _style_ingredients_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2, max_col=len(INGREDIENT_HEADERS)):
        for cell in row:
            if cell.column == 4:
                continue
            cell.font = INPUT_FONT
    rarity_validation = DataValidation(
        type="list",
        formula1='"common,uncommon,rare,epic,legendary"',
        allow_blank=False,
    )
    rarity_validation.add(f"H2:H{max(sheet.max_row, 200)}")
    sheet.add_data_validation(rarity_validation)
    shop_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    shop_validation.add(f"I2:I{max(sheet.max_row, 200)}")
    sheet.add_data_validation(shop_validation)
    widths = [18, 18, 20, 44, 12, 14, 10, 12, 14]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width


def _replace_sheet(
    wb,
    name: str,
    headers: list,
    rows: list[list],
    style_fn,
    index: int,
) -> None:
    if name in wb.sheetnames:
        del wb[name]
    sheet = wb.create_sheet(name, index)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    style_fn(sheet)


def _load_json_array(path: Path, label: str) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"Missing {label} file: {path}")
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError(f"{label} must be a JSON array.")
    return data


def sync_workbook_from_json() -> None:
    """Push runtime JSON content back into ingredients.xlsx."""
    if not XLSX_PATH.exists():
        create_default_workbook()

    ingredients = _load_json_array(INGREDIENTS_JSON, "ingredients")
    starter_bag = _load_json_array(STARTER_BAG_JSON, "starter bag")
    auras = _load_json_array(AURAS_JSON, "auras")
    trinkets = _load_json_array(TRINKETS_JSON, "trinkets")

    wb = load_workbook(XLSX_PATH)
    migrate_ingredients_sheet(wb)
    ensure_auras_sheet(wb)
    ensure_trinkets_sheet(wb)

    ingredient_rows = [
        [
            item["id"],
            item.get("art", item["id"]),
            item["display_name"],
            item["description"],
            item["point_value"],
            item["explosive_value"],
            item["shop_cost"],
            item["rarity"],
            bool(item.get("shop_available", True)),
        ]
        for item in ingredients
    ]
    starter_rows = [
        [item["id"], item["count"]]
        for item in starter_bag
    ]
    aura_rows = [
        [
            item["id"],
            item["display_name"],
            item["description"],
            item["pool"],
            item.get("pool_unlock_level", 1),
            item["explosion_limit_modifier"],
            item["score_multiplier_percent"],
            item["gold_multiplier_percent"],
        ]
        for item in auras
    ]
    trinket_rows = [
        [
            item["id"],
            item["display_name"],
            item["description"],
            bool(item.get("reward_offerable", True)),
        ]
        for item in trinkets
    ]

    _replace_sheet(wb, "Ingredients", INGREDIENT_HEADERS, ingredient_rows, _style_ingredients_sheet, 1)
    _replace_sheet(wb, "StarterBag", STARTER_HEADERS, starter_rows, _style_starter_bag_sheet, 2)
    _replace_sheet(wb, "Auras", AURA_HEADERS, aura_rows, _style_aura_sheet, 3)
    _replace_sheet(wb, "Trinkets", TRINKET_HEADERS, trinket_rows, _style_trinket_sheet, 4)

    try:
        wb.save(XLSX_PATH)
    except OSError as exc:
        raise OSError(
            f"Could not save spreadsheet ({exc}). "
            "Close ingredients.xlsx in Excel and run sync again."
        ) from exc

    print(f"Synced {len(ingredient_rows)} ingredients -> {XLSX_PATH}")
    print(f"Synced {len(starter_rows)} starter stacks")
    print(f"Synced {len(aura_rows)} auras")
    print(f"Synced {len(trinket_rows)} trinkets")


def _style_starter_bag_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.font = INPUT_FONT
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 10


def create_default_workbook() -> None:
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    readme = wb.active
    readme.title = "ReadMe"
    readme["A1"] = "Ingredient spreadsheet for Alchemy Roguelite"
    readme["A1"].font = Font(name="Arial", bold=True, size=14)
    readme["A3"] = "1. Edit the Ingredients, Auras, Trinkets, and StarterBag sheets."
    readme["A4"] = "2. Save this file."
    readme["A5"] = "3. Run: py tools/export_ingredients.py"
    readme["A6"] = "4. Play the game in Godot (it loads data/*.json)."
    readme["A8"] = "Ingredients: id, art, display_name, description, point_value, explosive_value, shop_cost, rarity, shop_available"
    readme["A9"] = "Auras: id, display_name, description, pool, explosion_limit_modifier, score_multiplier_percent, gold_multiplier_percent"
    readme["A10"] = "pool = normal (regular levels) or boss (every 5th level). Multipliers use 100 = no change."
    readme["A11"] = "art = PNG filename in assets/cards/ingredients/ (blank uses id)."
    readme["A12"] = "shop_available FALSE = starter-only, never appears in the shop."
    readme["A13"] = "Description: effect text (normal) or flavor text (italicize in Excel)."
    readme["A14"] = "Example art file: assets/cards/ingredients/boom_berry_1.png"
    readme["A15"] = "Trinkets: id, display_name, description, reward_offerable (passive run relics; art PNG optional in assets/cards/trinkets/)."
    readme["A16"] = "reward_offerable FALSE = never rolled as a boss trinket reward (can still be granted by effects/dev tools)."
    for row in range(3, 17):
        readme[f"A{row}"].font = NOTE_FONT
    readme.column_dimensions["A"].width = 78

    ingredients = wb.create_sheet("Ingredients")
    ingredients.append(INGREDIENT_HEADERS)
    for row in DEFAULT_INGREDIENTS:
        ingredients.append(row)
    for cell in ingredients[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in ingredients.iter_rows(min_row=2, max_col=len(INGREDIENT_HEADERS)):
        for cell in row:
            cell.font = INPUT_FONT
    rarity_validation = DataValidation(
        type="list",
        formula1='"common,uncommon,rare,epic,legendary"',
        allow_blank=False,
    )
    rarity_validation.add(f"G2:G{max(ingredients.max_row, 200)}")
    ingredients.add_data_validation(rarity_validation)
    shop_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    shop_validation.add(f"H2:H{max(ingredients.max_row, 200)}")
    ingredients.add_data_validation(shop_validation)
    widths = [18, 20, 44, 12, 14, 10, 12, 14]
    for idx, width in enumerate(widths, start=1):
        ingredients.column_dimensions[chr(64 + idx)].width = width

    starter = wb.create_sheet("StarterBag")
    starter.append(STARTER_HEADERS)
    for row in DEFAULT_STARTER_BAG:
        starter.append(row)
    for cell in starter[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in starter.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.font = INPUT_FONT
    starter.column_dimensions["A"].width = 20
    starter.column_dimensions["B"].width = 10

    auras = wb.create_sheet("Auras")
    auras.append(AURA_HEADERS)
    for row in DEFAULT_AURAS:
        auras.append(row)
    _style_aura_sheet(auras)

    trinkets = wb.create_sheet("Trinkets")
    trinkets.append(TRINKET_HEADERS)
    for row in DEFAULT_TRINKETS:
        trinkets.append(row)
    _style_trinket_sheet(trinkets)

    wb.save(XLSX_PATH)
    print(f"Created {XLSX_PATH}")


def _style_trinket_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2, max_col=len(TRINKET_HEADERS)):
        for cell in row:
            if cell.column == 3:
                continue
            cell.font = INPUT_FONT
    reward_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    reward_validation.add(f"D2:D{max(sheet.max_row, 200)}")
    sheet.add_data_validation(reward_validation)
    widths = [18, 22, 52, 16]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width


def ensure_trinkets_sheet(wb) -> bool:
    if "Trinkets" in wb.sheetnames:
        sheet = wb["Trinkets"]
        columns = _header_map(sheet)
        if set(TRINKET_HEADERS).issubset(columns):
            return False
        rows: list[list] = []
        for row_idx in range(2, sheet.max_row + 1):
            trinket_id = sheet.cell(row_idx, columns.get("id", 1)).value
            if trinket_id is None or str(trinket_id).strip() == "":
                continue
            display_name = (
                sheet.cell(row_idx, columns.get("display_name", columns.get("name", 2))).value
                if "display_name" in columns or "name" in columns
                else trinket_id
            )
            reward_offerable = True
            if "reward_offerable" in columns:
                reward_offerable = _parse_bool(
                    sheet.cell(row_idx, columns["reward_offerable"]).value,
                    default=True,
                )
            rows.append(
                [
                    str(trinket_id).strip(),
                    str(display_name or "").strip(),
                    _description_from_cell(sheet.cell(row_idx, columns.get("description", 3))),
                    reward_offerable,
                ]
            )
        if not rows:
            rows = [list(row) for row in DEFAULT_TRINKETS]
        del wb["Trinkets"]
        trinkets = wb.create_sheet("Trinkets")
        trinkets.append(TRINKET_HEADERS)
        for row in rows:
            trinkets.append(row)
        _style_trinket_sheet(trinkets)
        return True

    trinkets = wb.create_sheet("Trinkets")
    trinkets.append(TRINKET_HEADERS)
    for row in DEFAULT_TRINKETS:
        trinkets.append(row)
    _style_trinket_sheet(trinkets)
    return True


def _style_aura_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2, max_col=len(AURA_HEADERS)):
        for cell in row:
            cell.font = INPUT_FONT
    pool_validation = DataValidation(type="list", formula1='"normal,boss"', allow_blank=False)
    pool_validation.add(f"D2:D{max(sheet.max_row, 200)}")
    sheet.add_data_validation(pool_validation)
    widths = [18, 20, 44, 10, 18, 22, 22, 22]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width


def ensure_auras_sheet(wb) -> bool:
    if "Auras" in wb.sheetnames:
        sheet = wb["Auras"]
        columns = _header_map(sheet)
        if set(AURA_HEADERS).issubset(columns):
            return False
        rows: list[list] = []
        for row_idx in range(2, sheet.max_row + 1):
            aura_id = sheet.cell(row_idx, columns.get("id", 1)).value
            if aura_id is None or str(aura_id).strip() == "":
                continue
            rows.append(
                [
                    str(aura_id).strip(),
                    str(sheet.cell(row_idx, columns.get("display_name", 2)).value or "").strip(),
                    _description_from_cell(sheet.cell(row_idx, columns.get("description", 3))),
                    str(sheet.cell(row_idx, columns.get("pool", 4)).value or "normal").strip().lower(),
                    (
                        sheet.cell(row_idx, columns["pool_unlock_level"]).value
                        if "pool_unlock_level" in columns
                        else 1
                    ),
                    sheet.cell(row_idx, columns.get("explosion_limit_modifier", 6)).value or 0,
                    sheet.cell(row_idx, columns.get("score_multiplier_percent", 7)).value or 100,
                    sheet.cell(row_idx, columns.get("gold_multiplier_percent", 8)).value or 100,
                ]
            )
        if not rows:
            rows = [list(row) for row in DEFAULT_AURAS]
        del wb["Auras"]
        auras = wb.create_sheet("Auras")
        auras.append(AURA_HEADERS)
        for row in rows:
            auras.append(row)
        _style_aura_sheet(auras)
        return True

    auras = wb.create_sheet("Auras")
    auras.append(AURA_HEADERS)
    for row in DEFAULT_AURAS:
        auras.append(row)
    _style_aura_sheet(auras)
    return True


def migrate_starter_bag_sheet(wb, ingredient_ids: set[str]) -> bool:
    if "StarterBag" not in wb.sheetnames:
        return False
    sheet = wb["StarterBag"]
    columns = _header_map(sheet)
    if "ingredient_id" not in columns or "count" not in columns:
        return False

    kept: list[list] = []
    removed: list[str] = []
    for row_idx in range(2, sheet.max_row + 1):
        ingredient_id = sheet.cell(row_idx, columns["ingredient_id"]).value
        count = sheet.cell(row_idx, columns["count"]).value
        if ingredient_id is None or str(ingredient_id).strip() == "":
            continue
        ingredient_id = str(ingredient_id).strip()
        if ingredient_id not in ingredient_ids:
            removed.append(ingredient_id)
            continue
        kept.append([ingredient_id, count])

    if not removed and kept:
        return False

    if not kept:
        kept = [
            [row[0], row[1]]
            for row in DEFAULT_STARTER_BAG
            if row[0] in ingredient_ids
        ]

    del wb["StarterBag"]
    starter = wb.create_sheet("StarterBag")
    starter.append(STARTER_HEADERS)
    for row in kept:
        starter.append(row)
    for cell in starter[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in starter.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.font = INPUT_FONT
    starter.column_dimensions["A"].width = 20
    starter.column_dimensions["B"].width = 10
    try:
        wb.save(XLSX_PATH)
        if removed:
            print(f"Removed stale starter bag entries: {', '.join(removed)}")
        print(f"Updated StarterBag sheet in {XLSX_PATH}")
    except OSError as exc:
        print(
            f"Could not save migrated StarterBag sheet ({exc}). "
            "Close ingredients.xlsx in Excel and export again to update the file."
        )
    return True


def export_workbook() -> None:
    if not XLSX_PATH.exists():
        create_default_workbook()

    wb = load_workbook(XLSX_PATH)
    migrate_ingredients_sheet(wb)
    auras_added = ensure_auras_sheet(wb)
    trinkets_added = ensure_trinkets_sheet(wb)
    if auras_added or trinkets_added:
        try:
            wb.save(XLSX_PATH)
            if auras_added:
                print(f"Added Auras sheet to {XLSX_PATH}")
            if trinkets_added:
                print(f"Added Trinkets sheet to {XLSX_PATH}")
        except OSError as exc:
            print(
                f"Could not save spreadsheet sheets ({exc}). "
                "Close ingredients.xlsx in Excel and export again to update the file."
            )

    if "Ingredients" not in wb.sheetnames:
        raise ValueError("Missing 'Ingredients' sheet in ingredients.xlsx")
    if "Auras" not in wb.sheetnames:
        raise ValueError("Missing 'Auras' sheet in ingredients.xlsx")
    if "StarterBag" not in wb.sheetnames:
        raise ValueError("Missing 'StarterBag' sheet in ingredients.xlsx")
    if "Trinkets" not in wb.sheetnames:
        raise ValueError("Missing 'Trinkets' sheet in ingredients.xlsx")

    sheet = wb["Ingredients"]
    columns = _header_map(sheet)
    missing = set(INGREDIENT_HEADERS) - set(columns)
    if missing:
        raise ValueError(f"Ingredients sheet missing columns: {sorted(missing)}")

    ingredients = []
    seen_ids: set[str] = set()
    for row_index in range(2, sheet.max_row + 1):
        ingredient_id = sheet.cell(row_index, columns["id"]).value
        if ingredient_id is None or str(ingredient_id).strip() == "":
            continue
        ingredient_id = str(ingredient_id).strip()
        if ingredient_id in seen_ids:
            raise ValueError(f"Duplicate ingredient id '{ingredient_id}' on row {row_index}.")
        seen_ids.add(ingredient_id)

    migrate_starter_bag_sheet(wb, seen_ids)

    for row_index in range(2, sheet.max_row + 1):
        ingredient_id = sheet.cell(row_index, columns["id"]).value
        if ingredient_id is None or str(ingredient_id).strip() == "":
            continue
        ingredient_id = str(ingredient_id).strip()

        description = _description_from_cell(sheet.cell(row_index, columns["description"]))
        rarity = str(sheet.cell(row_index, columns["rarity"]).value).strip().lower()
        art = _normalize_art(sheet.cell(row_index, columns["art"]).value, ingredient_id)
        ingredients.append(
            {
                "id": ingredient_id,
                "art": art,
                "display_name": str(sheet.cell(row_index, columns["display_name"]).value).strip(),
                "description": description,
                "point_value": _parse_int(
                    sheet.cell(row_index, columns["point_value"]).value,
                    "point_value",
                    row_index,
                    ingredient_id,
                ),
                "explosive_value": _parse_int_default(
                    sheet.cell(row_index, columns["explosive_value"]).value,
                    "explosive_value",
                    row_index,
                    ingredient_id,
                ),
                "shop_cost": _parse_int(
                    sheet.cell(row_index, columns["shop_cost"]).value,
                    "shop_cost",
                    row_index,
                    ingredient_id,
                ),
                "rarity": rarity,
                "shop_available": _parse_bool(
                    sheet.cell(row_index, columns["shop_available"]).value,
                    ingredient_id,
                ),
            }
        )

    starter_sheet = wb["StarterBag"]
    starter_rows = _sheet_rows_values(starter_sheet)
    starter_header = [str(cell).strip() for cell in starter_rows[0]]
    if starter_header != STARTER_HEADERS:
        raise ValueError("StarterBag headers must be: ingredient_id, count")

    starter_bag = []
    for row_index, row in enumerate(starter_rows[1:], start=2):
        values = list(row) + [None, None]
        row_ingredient_id = str(values[0]).strip() if values[0] is not None else ""
        if row_ingredient_id == "":
            continue
        if row_ingredient_id not in seen_ids:
            raise ValueError(
                f"StarterBag row {row_index}: unknown ingredient_id '{row_ingredient_id}'."
            )
        count = _parse_int(values[1], "count", row_index, row_ingredient_id)
        if count <= 0:
            raise ValueError(f"StarterBag row {row_index}: count must be > 0.")
        starter_bag.append({"id": row_ingredient_id, "count": count})

    aura_sheet = wb["Auras"]
    aura_columns = _header_map(aura_sheet)
    missing_aura_columns = set(AURA_HEADERS) - set(aura_columns)
    if missing_aura_columns:
        raise ValueError(f"Auras sheet missing columns: {sorted(missing_aura_columns)}")

    auras = []
    seen_aura_ids: set[str] = set()
    for row_index in range(2, aura_sheet.max_row + 1):
        aura_id = aura_sheet.cell(row_index, aura_columns["id"]).value
        if aura_id is None or str(aura_id).strip() == "":
            continue
        aura_id = str(aura_id).strip()
        if aura_id in seen_aura_ids:
            raise ValueError(f"Duplicate aura id '{aura_id}' on row {row_index}.")
        seen_aura_ids.add(aura_id)

        pool = str(aura_sheet.cell(row_index, aura_columns["pool"]).value).strip().lower()
        if pool not in {"normal", "boss"}:
            raise ValueError(f"Auras row {row_index} ('{aura_id}'): pool must be 'normal' or 'boss'.")
        pool_unlock_level = _parse_int(
            aura_sheet.cell(row_index, aura_columns["pool_unlock_level"]).value,
            "pool_unlock_level",
            row_index,
            aura_id,
        )
        if pool_unlock_level < 1:
            raise ValueError(
                f"Auras row {row_index} ('{aura_id}'): pool_unlock_level must be >= 1."
            )
        explosion_mod = _parse_int(
            aura_sheet.cell(row_index, aura_columns["explosion_limit_modifier"]).value,
            "explosion_limit_modifier",
            row_index,
            aura_id,
        )
        score_mult = _parse_int(
            aura_sheet.cell(row_index, aura_columns["score_multiplier_percent"]).value,
            "score_multiplier_percent",
            row_index,
            aura_id,
        )
        gold_mult = _parse_int(
            aura_sheet.cell(row_index, aura_columns["gold_multiplier_percent"]).value,
            "gold_multiplier_percent",
            row_index,
            aura_id,
        )
        auras.append(
            {
                "id": aura_id,
                "display_name": str(aura_sheet.cell(row_index, aura_columns["display_name"]).value).strip(),
                "description": _description_from_cell(aura_sheet.cell(row_index, aura_columns["description"])),
                "pool": pool,
                "pool_unlock_level": pool_unlock_level,
                "explosion_limit_modifier": explosion_mod,
                "score_multiplier_percent": score_mult,
                "gold_multiplier_percent": gold_mult,
            }
        )

    if not auras:
        raise ValueError("Auras sheet has no aura rows.")

    trinket_sheet = wb["Trinkets"]
    trinket_columns = _header_map(trinket_sheet)
    missing_trinket_columns = set(TRINKET_HEADERS) - set(trinket_columns)
    if missing_trinket_columns:
        raise ValueError(f"Trinkets sheet missing columns: {sorted(missing_trinket_columns)}")

    trinkets = []
    seen_trinket_ids: set[str] = set()
    for row_index in range(2, trinket_sheet.max_row + 1):
        trinket_id = trinket_sheet.cell(row_index, trinket_columns["id"]).value
        if trinket_id is None or str(trinket_id).strip() == "":
            continue
        trinket_id = str(trinket_id).strip()
        if trinket_id in seen_trinket_ids:
            raise ValueError(f"Duplicate trinket id '{trinket_id}' on row {row_index}.")
        seen_trinket_ids.add(trinket_id)
        trinkets.append(
            {
                "id": trinket_id,
                "display_name": str(
                    trinket_sheet.cell(row_index, trinket_columns["display_name"]).value
                ).strip(),
                "description": _description_from_cell(
                    trinket_sheet.cell(row_index, trinket_columns["description"])
                ),
                "reward_offerable": _parse_bool(
                    trinket_sheet.cell(
                        row_index, trinket_columns["reward_offerable"]
                    ).value,
                    default=True,
                ),
            }
        )

    INGREDIENTS_JSON.write_text(json.dumps(ingredients, indent=2) + "\n", encoding="utf-8")
    STARTER_BAG_JSON.write_text(json.dumps(starter_bag, indent=2) + "\n", encoding="utf-8")
    AURAS_JSON.write_text(json.dumps(auras, indent=2) + "\n", encoding="utf-8")
    TRINKETS_JSON.write_text(json.dumps(trinkets, indent=2) + "\n", encoding="utf-8")
    print(f"Exported {len(ingredients)} ingredients -> {INGREDIENTS_JSON}")
    print(f"Exported {len(auras)} auras -> {AURAS_JSON}")
    print(f"Exported {len(trinkets)} trinkets -> {TRINKETS_JSON}")
    print(f"Exported {len(starter_bag)} starter stacks -> {STARTER_BAG_JSON}")


def main() -> int:
    if len(sys.argv) > 1 and sys.argv[1] == "--init":
        create_default_workbook()
        export_workbook()
        return 0
    if len(sys.argv) > 1 and sys.argv[1] == "--sync-json":
        try:
            sync_workbook_from_json()
        except Exception as exc:
            print(f"Sync failed: {exc}", file=sys.stderr)
            return 1
        return 0
    try:
        export_workbook()
    except Exception as exc:
        print(f"Export failed: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())